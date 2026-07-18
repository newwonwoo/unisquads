"""
IROS 등기고유번호 리졸버 (자기적응형 미니 RPA)
================================================
addr-refine 정제주소 → 인터넷등기소 부동산 열람 검색 → 등기고유번호(14자리) 추출

[실화면 근거 — 2026-07 확인]
- 부동산 등기사항증명서 열람·발급 신청 > 간편검색 탭
- 검색 단계는 비회원 접근 가능(캡차·로그인 없음). 결제 단계만 전화번호+비번.
- 필수 입력 3: 주소 / 부동산구분(기본 전체) / 시/도
- 도로명 병기 안 된 등기는 도로명 검색 불가 → **지번주소** 사용 원칙
- 결과: 테이블 행마다 [번호|부동산고유번호|부동산구분|부동산표시|등기상태]
  한 주소에 토지·건물 등 복수 물건이 정상 → MULTIPLE 기본
- '다음'/'결제대상목록'이 결제 경로 → 절대 클릭 금지(0원 보장)

[바이브코더용 설명]
- '무인 봇'이 아니라 '사람이 실행 버튼을 누르는 반자동 매크로'. 창이 눈앞에 뜸.
- 검색까지만 자동, 결제는 절대 안 건드림.
- 셀렉터를 하드코딩하지 않고, 실행 시 화면을 훑어 입력칸·버튼·시/도 드롭다운을 찾음.
- 고유번호는 형식(4-4-6 = 14자리)이 불변이라, 화면이 바뀌어도 정규식으로 잡힘.

[사용 전제]
- 로컬 PC 또는 EC2 실행 (Playwright + 브라우저). 폰 불가.
- pip install playwright && playwright install chromium
- 저빈도 개인 업무 보조 용도. 대량/무인/상시 자동화 아님.
"""

import re
import sys
import time
import argparse
from dataclasses import dataclass, field
from typing import Optional

# 고유번호 형식: 4자리-4자리-6자리 (구분자 유무 모두 허용)
UNIQUE_NO_RE = re.compile(r"\b(\d{4})\s*-\s*(\d{4})\s*-\s*(\d{6})\b")

IROS_ENTRY = "https://www.iros.go.kr"

# 검색 입력칸을 찾을 때 참고할 키워드
ADDR_INPUT_HINTS = ["주소", "도로명", "소재지", "지번", "addr", "juso"]
SEARCH_BTN_HINTS = ["검색"]
# 결제/발급/다음 단계로 넘어가는 위험 버튼 — 절대 클릭 금지 (0원 보장 가드)
FORBIDDEN_HINTS = ["결제", "발급", "열람하기", "결제하기", "출력", "다음", "선택완료", "결제대상"]

# 시/도 명칭 정규화 (정제주소 앞토큰 → IROS 드롭다운 표기)
SIDO_MAP = {
    "서울": "서울특별시", "서울시": "서울특별시", "서울특별시": "서울특별시",
    "부산": "부산광역시", "부산시": "부산광역시", "부산광역시": "부산광역시",
    "대구": "대구광역시", "대구광역시": "대구광역시",
    "인천": "인천광역시", "인천광역시": "인천광역시",
    "광주": "광주광역시", "광주광역시": "광주광역시",
    "대전": "대전광역시", "대전광역시": "대전광역시",
    "울산": "울산광역시", "울산광역시": "울산광역시",
    "세종": "세종특별자치시", "세종시": "세종특별자치시", "세종특별자치시": "세종특별자치시",
    "경기": "경기도", "경기도": "경기도",
    "강원": "강원특별자치도", "강원도": "강원특별자치도", "강원특별자치도": "강원특별자치도",
    "충북": "충청북도", "충청북도": "충청북도",
    "충남": "충청남도", "충청남도": "충청남도",
    "전북": "전북특별자치도", "전라북도": "전북특별자치도", "전북특별자치도": "전북특별자치도",
    "전남": "전라남도", "전라남도": "전라남도",
    "경북": "경상북도", "경상북도": "경상북도",
    "경남": "경상남도", "경상남도": "경상남도",
    "제주": "제주특별자치도", "제주도": "제주특별자치도", "제주특별자치도": "제주특별자치도",
}


@dataclass
class ResolveResult:
    address: str
    status: str                      # RESOLVED | MULTIPLE | NOT_FOUND | BLOCKED | ERROR
    unique_no: Optional[str] = None  # 단건 확정 시
    candidates: list = field(default_factory=list)  # 최종 판정 후보
    message: str = ""
    all_candidates: list = field(default_factory=list)  # PNU/지번 전체수집 후보
    complete: bool = False
    total_count: Optional[int] = None
    received_count: int = 0
    pages_fetched: int = 0
    strategy: str = ""
    parser_version: str = "iros-parser-v2"


def normalize_sido(address: str) -> Optional[str]:
    """정제주소 앞부분에서 시/도를 뽑아 IROS 드롭다운 표기로 변환."""
    first = address.strip().split()[0] if address.strip() else ""
    # 앞토큰 직접 매칭
    if first in SIDO_MAP:
        return SIDO_MAP[first]
    # 앞토큰이 '서울특별시' 등 완전형인데 맵에 없을 때 대비: 부분 매칭
    for key, val in SIDO_MAP.items():
        if address.startswith(key):
            return val
    return None


def strip_sido(address: str, sido_full: Optional[str]) -> str:
    """주소 입력칸에는 시/도를 뺀 나머지를 넣는다(드롭다운에서 시/도 별도 선택하므로).
    IROS placeholder가 '서초구 서초동 967' 형태인 점 반영."""
    if not sido_full:
        return address
    # 앞의 시/도 토큰 제거 (약칭·완전형 모두)
    tokens = address.strip().split()
    if tokens and (tokens[0] in SIDO_MAP or tokens[0] == sido_full):
        return " ".join(tokens[1:])
    return address


def _normalize_unique(m: re.Match) -> str:
    return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"


def _all_frames(page):
    """메인 + 모든 iframe을 훑기 위해 프레임 목록 반환.
    IROS는 프레임 기반 구조(excelUploadFrame__, ozReportFrame__ 등)라
    검색 입력·결과가 특정 iframe에 있을 수 있음. 전 프레임 순회가 필수."""
    try:
        return list(page.frames)
    except Exception:
        return [page]


def _find_address_input(page):
    """모든 프레임의 input을 훑어 '주소 입력칸'을 힌트 점수제로 탐색.
    반환: (element, frame) — 시/도·검색을 같은 프레임에서 처리하기 위함."""
    best, best_score, best_frame = None, -1, None
    for fr in _all_frames(page):
        try:
            inputs = fr.query_selector_all("input[type='text'], input:not([type])")
        except Exception:
            continue
        for el in inputs:
            try:
                if not el.is_visible():
                    continue
                attrs = " ".join(filter(None, [
                    el.get_attribute("placeholder") or "",
                    el.get_attribute("name") or "",
                    el.get_attribute("id") or "",
                    el.get_attribute("title") or "",
                    el.get_attribute("aria-label") or "",
                ])).lower()
            except Exception:
                continue
            score = sum(1 for h in ADDR_INPUT_HINTS if h.lower() in attrs)
            if score > best_score:
                best_score, best, best_frame = score, el, fr
    return (best, best_frame) if best_score > 0 else (None, None)


def _select_sido(frame, sido_full: str) -> bool:
    """주어진 프레임(또는 전 프레임)에서 시/도 <select>를 찾아 선택."""
    frames = [frame] if frame is not None else []
    for sel_root in frames:
        try:
            for sel in sel_root.query_selector_all("select"):
                if not sel.is_visible():
                    continue
                options = [o.inner_text().strip() for o in sel.query_selector_all("option")]
                if sido_full in options:
                    sel.select_option(label=sido_full)
                    return True
        except Exception:
            continue
    return False


def _click_search(page, frame=None) -> bool:
    """검색 버튼 탐색 후 클릭. 입력칸이 속한 프레임 우선, 없으면 전 프레임.
    위험 버튼(결제/발급/다음)은 회피."""
    search_scopes = [frame] if frame is not None else []
    search_scopes += [f for f in _all_frames(page) if f is not frame]
    for scope in search_scopes:
        if scope is None:
            continue
        try:
            elems = scope.query_selector_all("button, a, input[type='button'], input[type='submit']")
        except Exception:
            continue
        for el in elems:
            try:
                tag = el.evaluate("e => e.tagName")
                label = (el.get_attribute("value") if tag == "INPUT" else el.inner_text()) or ""
                label = label.strip()
                if not label or not el.is_visible():
                    continue
                if any(f in label for f in FORBIDDEN_HINTS):
                    continue
                if any(h in label for h in SEARCH_BTN_HINTS):
                    el.click()
                    return True
            except Exception:
                continue
    return False


def _parse_result_rows(page):
    """모든 프레임의 결과 테이블을 행 단위로 파싱.
    결과가 어느 iframe에 렌더링되든 잡아냄. 실패 시 전 프레임 텍스트 정규식 폴백."""
    out, seen = [], set()
    for fr in _all_frames(page):
        try:
            rows = fr.query_selector_all("table tbody tr, table tr")
        except Exception:
            continue
        for tr in rows:
            try:
                txt = tr.inner_text()
            except Exception:
                continue
            m = UNIQUE_NO_RE.search(txt)
            if not m:
                continue
            uno = _normalize_unique(m)
            if uno in seen:
                continue
            seen.add(uno)
            cells = [c.strip() for c in re.split(r"[\t\n]", txt) if c.strip()]
            gubun = next((c for c in cells if c in ("토지", "건물", "집합건물")), "")
            state = next((c for c in cells if c in ("현행", "폐쇄")), "")
            sojae = next((c for c in cells if any(s in c for s in
                         ["특별시", "광역시", "특별자치", "도 "]) or "시 " in c), "")
            out.append({"unique_no": uno, "gubun": gubun, "sojae": sojae, "state": state})
    # 폴백: 전 프레임 body 텍스트에서 번호만
    if not out:
        for fr in _all_frames(page):
            try:
                body = fr.inner_text("body")
            except Exception:
                continue
            for m in UNIQUE_NO_RE.finditer(body):
                uno = _normalize_unique(m)
                if uno not in seen:
                    seen.add(uno)
                    out.append({"unique_no": uno, "gubun": "", "sojae": "", "state": ""})
    return out


def _detect_block(page) -> bool:
    """전 프레임에서 캡차/로그인 감지."""
    for fr in _all_frames(page):
        try:
            txt = fr.inner_text("body")
        except Exception:
            continue
        for kw in ["자동입력 방지", "보안문자", "캡차", "captcha", "로그인이 필요"]:
            if kw.lower() in txt.lower():
                return True
    return False


def resolve_one(page, address: str, wait_human: float = 0.0) -> ResolveResult:
    """단일 주소 → 고유번호. page 재사용. iframe 다중 구조 대응."""
    try:
        page.goto(IROS_ENTRY, wait_until="domcontentloaded", timeout=25000)
        time.sleep(1.5)  # 프레임·보안모듈 로딩 여유

        sido_full = normalize_sido(address)
        addr_for_input = strip_sido(address, sido_full)

        inp, frame = _find_address_input(page)
        if inp is None:
            return ResolveResult(address, "ERROR",
                                 message="주소 입력칸을 찾지 못했습니다(프레임 구조/화면 변경 가능). 수동 확인 필요.")
        inp.click()
        inp.fill("")
        inp.type(addr_for_input, delay=15)
        time.sleep(0.3)

        # 시/도 드롭다운 (입력칸과 같은 프레임 우선, 없으면 전 프레임)
        if sido_full:
            if not _select_sido(frame, sido_full):
                _select_sido(page.main_frame, sido_full)
            time.sleep(0.2)

        if not _click_search(page, frame):
            inp.press("Enter")
        time.sleep(2.0)  # 결과 렌더링(프레임) 대기

        if _detect_block(page):
            if wait_human > 0:
                time.sleep(wait_human)
            else:
                return ResolveResult(address, "BLOCKED",
                                     message="로그인/보안문자 감지. 창에서 직접 처리 후 재시도.")

        cands = _parse_result_rows(page)
        if len(cands) == 0:
            return ResolveResult(address, "NOT_FOUND",
                                 message="고유번호 없음. 주소 정밀도/미등록/도로명전용 여부 확인(지번주소 권장).")
        if len(cands) == 1:
            c = cands[0]
            desc = " ".join(filter(None, [c["gubun"], c["state"]]))
            return ResolveResult(address, "RESOLVED", unique_no=c["unique_no"],
                                 candidates=cands, message=desc or c["sojae"])
        return ResolveResult(address, "MULTIPLE", candidates=cands,
                             message=f"{len(cands)}건(토지/건물 등). 부동산구분으로 구별 필요.")
    except Exception as e:
        return ResolveResult(address, "ERROR", message=f"{type(e).__name__}: {e}")


def resolve_batch(addresses, headless=False, throttle=2.0, wait_human=0.0):
    """주소 리스트 → 결과 리스트. 브라우저 1회 기동, 요청 간 지연."""
    from playwright.sync_api import sync_playwright
    results = []
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        page = browser.new_page()
        page.set_default_timeout(15000)
        for i, addr in enumerate(addresses):
            r = resolve_one(page, addr, wait_human=wait_human)
            results.append(r)
            print(f"[{i+1}/{len(addresses)}] {addr} → {r.status}"
                  + (f" {r.unique_no}" if r.unique_no else ""), file=sys.stderr)
            if i < len(addresses) - 1:
                time.sleep(throttle)
        browser.close()
    return results


def main():
    ap = argparse.ArgumentParser(description="IROS 등기고유번호 리졸버 (자기적응형 미니 RPA)")
    ap.add_argument("--input", help="주소 목록 파일(줄바꿈) 또는 xlsx(A열)")
    ap.add_argument("--addr", help="단일 주소 직접 입력")
    ap.add_argument("--out", default="iros-결과.xlsx", help="결과 xlsx 경로")
    ap.add_argument("--headless", action="store_true", help="창 숨김(EC2/서버는 필수)")
    ap.add_argument("--throttle", type=float, default=2.0, help="요청 간 지연(초)")
    ap.add_argument("--wait-human", type=float, default=0.0, help="캡차/로그인 시 대기(초)")
    args = ap.parse_args()

    addresses = []
    if args.addr:
        addresses = [args.addr]
    elif args.input:
        if args.input.lower().endswith((".xlsx", ".xls")):
            import openpyxl
            wb = openpyxl.load_workbook(args.input)
            ws = wb.active
            for row in ws.iter_rows(min_row=1, values_only=True):
                v = str(row[0] or "").strip() if row else ""
                if v and not re.search(r"주소|address", v, re.I):
                    addresses.append(v)
        else:
            with open(args.input, encoding="utf-8") as f:
                addresses = [ln.strip() for ln in f if ln.strip()]
    else:
        print("입력이 없습니다. --addr 또는 --input 사용", file=sys.stderr)
        sys.exit(1)

    results = resolve_batch(addresses, headless=args.headless,
                            throttle=args.throttle, wait_human=args.wait_human)
    write_xlsx(results, args.out)
    print(f"\n완료: {args.out} ({len(results)}건)", file=sys.stderr)


def write_xlsx(results, path):
    """결과 → 엑셀. 고유번호 텍스트 서식(하이픈 보존)."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "등기고유번호"
    head = ["주소", "상태", "등기고유번호", "부동산구분", "등기상태", "소재지/비고", "후보수", "후보상세"]
    ws.append(head)
    for c in ws[1]:
        c.font = Font(bold=True)
        c.fill = PatternFill("solid", start_color="E4EFE9")
    for r in results:
        first = r.candidates[0] if r.candidates else {}
        cand_detail = " | ".join(
            f'{c["unique_no"]}({c.get("gubun","")}·{c.get("state","")})' for c in r.candidates
        ) if r.candidates else ""
        ws.append([
            r.address, r.status, r.unique_no or "",
            first.get("gubun", ""), first.get("state", ""),
            first.get("sojae", "") or r.message,
            len(r.candidates) if r.candidates else 0,
            cand_detail,
        ])
    for row in ws.iter_rows(min_row=2, min_col=3, max_col=3):
        for cell in row:
            cell.number_format = "@"
    for col, w in zip("ABCDEFGH", [34, 11, 18, 12, 10, 40, 8, 56]):
        ws.column_dimensions[col].width = w
    wb.save(path)


if __name__ == "__main__":
    main()
